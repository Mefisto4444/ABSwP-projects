import openpyxl
from openpyxl.utils import get_column_letter

print("Otwieranie skoroszytu....")
wb = openpyxl.load_workbook("Moving.xlsx")
sheet = wb.active
sheetData = {}

print("Pobieranie danych i czyszczenie skoroszytu....")
for col in range(1,sheet.max_column+1):
    for row in range(1,sheet.max_row+1):
        sheetData.update({(col,row):sheet.cell(row,col).value})
        sheet.cell(row,col).value = None

print("Wprowadzanie danych do skoroszytu....")    
for key in sheetData.keys():
    sheet[f"{get_column_letter(key[1])}{key[0]}"].value = sheetData[key]

print("Zapisywanie skoroszytu do pliku....")
try:
    wb.save("Moving.xlsx")
except PermissionError:
    print("Wyłącz podgląd pliku.")