import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
    n = sys.argv[1]
else:
    n = int(input(": "))

wb = openpyxl.Workbook()
sheet = wb[wb.sheetnames[0]]
BoldFontObj = Font(bold=True)

for i in range(1,n+1):
    sheet.cell(i+1,1).font = BoldFontObj
    sheet.cell(i+1,1).value = i
    sheet.cell(1,i+1).font = BoldFontObj
    sheet.cell(1,i+1).value = i

for col_numb in range(2,sheet.max_column+1):
    for row_numb in range(2,sheet.max_row+1):
        sheet[f"{get_column_letter(col_numb)}{row_numb}"].value = sheet.cell(1,col_numb).value * sheet.cell(row_numb,1).value

try:
    wb.save("Tabliczka.xlsx")
except PermissionError:
    print("Zamknij podgląd pliku.")