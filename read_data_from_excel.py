import openpyxl, pprint

counties = {}

print("Otwieranie skoroszytu....")
wb = openpyxl.load_workbook("censuspopdata.xlsx")
sheet = wb[wb.sheetnames[0]]

print("Uporządkowywanie danych....")
for i in range(2,sheet.max_row):
    if sheet.cell(i,3).value not in counties:
        counties[sheet.cell(i,3).value] = sheet.cell(i,4).value
    else:
        counties[sheet.cell(i,3).value] = counties[sheet.cell(i,3).value] + sheet.cell(i,4).value

print("Zapisywanie danych do plików....")
with open("censuspopdata.txt",'w') as cFile:
    cFile.write(pprint.pformat(counties))
with open("census2010.py",'w') as cpFile:
    cpFile.write("census_data = "+pprint.pformat(counties))

print("Zakończono proces.")