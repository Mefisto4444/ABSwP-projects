import openpyxl, sys
from openpyxl.utils import get_column_letter

def Blank(name:str,start:int,M:int):
    wb = openpyxl.load_workbook(name)
    sheet = wb.active
    
    for col_num in range(1,sheet.max_column+1):
        for row_num in range(start,start +M):
            sheet[f"{get_column_letter(col_num)}{row_num}"].value = None

    try:
        wb.save(name)
    except PermissionError:
        print("Zamknij podgląd pliku.")

if len(sys.argv) > 1:
    Blank(sys.argv[1], sys.argv[2], sys.argv[3])
else:
    Blank("produceSales.xlsx",10,2)