import openpyxl

to_update = {}
try:
    while True:
        to_update.update({input("Produkt: "):round(float(input("Cena: ")),2)})
except KeyboardInterrupt:
    print("Zakończono uaktaulnianie bazy wstępnej.")
print(to_update)

print("Otwieranie skoroszytu....")
wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb[wb.sheetnames[0]]

print("Wprowadzanie zmian w skoroszycie....")
for i in range(2,sheet.max_row):
    if sheet.cell(i,1).value in to_update.keys():
        sheet.cell(i,2).value = to_update[sheet.cell(i,1).value]

print("Zapisywanie skoroszytu....")
wb.save("produceSales.xlsx")
print("Zakończono program.")