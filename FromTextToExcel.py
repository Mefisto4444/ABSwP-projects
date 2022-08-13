import os, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FILE_KATALOG = "D:\\jakub\\reddit\\files_for_excel\\"

wb = openpyxl.Workbook()
sheet = wb.active
BoldTxt = Font(bold=True)

sheet["A1"].font = BoldTxt
sheet["A1"].value = "Imie"
sheet["B1"].font = BoldTxt
sheet["B1"].value = "Stanowisko"
sheet["C1"].font = BoldTxt
sheet["C1"].value = "Zalega"
sheet.freeze_panes = "A2"

data_list = []
for file in os.listdir(FILE_KATALOG):
    with open(f"{FILE_KATALOG}{file}",'r',encoding="utf-8") as File:
        data_list.append(File.readlines())
        File.close()

for file_num in range(len(data_list)):
    line = 2
    for row in data_list[file_num]:
        sheet[f"{get_column_letter(file_num+1)}{line}"].value = row
        line = line + 1
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 15

try:
    wb.save("FromTextToExcel.xlsx")
except PermissionError:
    print("Wyłącz podgląd pliku.")