import openpyxl
from openpyxl.utils import get_column_letter
PATH = "D:\\jakub\\reddit\\files_for_excel2\\"

wb = openpyxl.load_workbook("FromExcelToText.xlsx")
sheet = wb.active

for col in range(1,sheet.max_column+1):
    with open(f"{PATH}file{col}.txt",'w') as File:
        for row in range(1,sheet.max_row+1):
            File.write(f'{sheet[f"{get_column_letter(col)}{row}"].value}')
